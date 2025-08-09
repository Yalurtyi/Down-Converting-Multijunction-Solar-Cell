#GitHub UC-DC-MJ Bandgap Optimizer
import pandas as pd
import numpy as np
from scipy.constants import h, c, e, k
from scipy import integrate
from scipy.optimize import differential_evolution
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Alignment
file_path = r"D:\dowloads\AM1.5Global.xlsx" #Import AM1.5G spectrum here
OUTPUT_FILE = r"D:\dowloads\ucdcmj_results.xlsx" #Enter desired data download path
data = pd.read_excel(file_path)
wavelength_np = data.iloc[:, 0].values    # nm
irradiance_np = data.iloc[:, 1].values    # W/m^2/nm
q = e
T = 300
pi = np.pi
n = 1
CM = 1
def compute_j0(Eg):
    prefactor = 2 * pi * q**4 / (h**3 * c**2)
    def integrand(E):
        x = np.clip(q * E / (k * T), -700, 700)
        return E**2 / (np.expm1(x))
    try:
        integral, _ = integrate.quad(integrand, Eg, np.inf, limit=200)
    except Exception:
        return np.inf
    return prefactor * integral
def model_eff(dc_Eg, uc_Eg, mj_Eg, irradiance_np, wavelength_np):
    E_spectrum = h * c / (q * wavelength_np * 1e-9)  # Energy in eV
    I_processed = irradiance_np.copy()
    P_in = np.trapz(I_processed, wavelength_np)  # total incident power
    dλ = np.gradient(wavelength_np)

    # Down-conversion
    for Eg_target in dc_Eg:
        indices_src = np.where(np.abs(E_spectrum - Eg_target) <= 0.05)[0]
        for i in indices_src:
            if I_processed[i] <= 0:
                continue
            E_tgt = E_spectrum[i] / 2.0
            indices_dst = np.where(np.abs(E_spectrum - E_tgt) <= 0.05)[0]
            if len(indices_dst) == 0:
                continue
            denom = np.sum(dλ[indices_dst])
            if denom <= 0:
                continue
            I_processed[indices_dst] += I_processed[i] * dλ[i] / denom
            I_processed[i] = 0.0

    # Up-conversion
    for Eg_target in uc_Eg:
        indices_src = np.where(np.abs(E_spectrum - Eg_target) <= 0.05)[0]
        for i in indices_src:
            if I_processed[i] <= 0:
                continue
            E_uc = 2 * E_spectrum[i]
            j = np.argmin(np.abs(E_spectrum - E_uc))
            if np.abs(E_spectrum[j] - E_uc) > 0.05:
                continue
            # available photons in source bin
            available_photons = I_processed[i] * dλ[i] / (E_spectrum[i] * q)
            conversions = available_photons / 2.0
            # energy bookkeeping (same approach you used)
            energy_removed = 2 * conversions * E_spectrum[i] * q
            energy_added = conversions * E_uc * q
            # avoid negative due to rounding
            I_processed[i] = max(0.0, I_processed[i] - energy_removed / dλ[i])
            I_processed[j] += energy_added / dλ[j]

    # Multijunction cell calculation
    P_max_list = []
    for i, Eg_i in enumerate(mj_Eg):
        if i == 0:
            λ_max = h * c / (q * 1e-9 * Eg_i)
            mask = wavelength_np <= λ_max
        else:
            λ_min = h * c / (q * 1e-9 * mj_Eg[i-1])
            λ_max = h * c / (q * 1e-9 * Eg_i)
            mask = (wavelength_np >= λ_min) & (wavelength_np <= λ_max)

        if not np.any(mask):
            P_max_list.append(0.0)
            continue

        E_cell = h * c / (q * wavelength_np[mask] * 1e-9)
        photon_flux = np.trapz(I_processed[mask] / (q * E_cell), wavelength_np[mask])
        J_ph = CM * q * photon_flux
        J0 = compute_j0(Eg_i)

        if J_ph <= 0 or J0 <= 0 or not np.isfinite(J0):
            P_max_list.append(0.0)
            continue

        Voc = n * k * T * np.log1p(J_ph / J0) / q
        V = np.linspace(0, Voc, 500)
        J_total = J_ph - J0 * np.expm1(V * q / (n * k * T))
        power = J_total * V
        max_power = np.max(power) if len(power) > 0 else 0.0
        P_max_list.append(max_power)

    P_collected = sum(P_max_list)
    if P_in <= 0:
        return 0.0
    return (P_collected / P_in) * 100.0
def find_optimal_configuration(dc_config, uc_config, mj_config):
    configs = [(n_dc, n_uc, n_mj) for n_dc in dc_config for n_uc in uc_config for n_mj in mj_config]
    results = []
    print(f"Testing {len(configs)} configurations...")
    for idx, (n_dc, n_uc, n_mj) in enumerate(configs, start=1):
        n_params = n_dc + n_uc + n_mj
        bounds = [(0.1, 5.0)] * n_params
        def objective(x):
            dc_Eg = sorted(x[:n_dc], reverse=True) if n_dc > 0 else []
            uc_Eg = list(x[n_dc:n_dc+n_uc]) if n_uc > 0 else []
            mj_Eg = sorted(x[n_dc+n_uc:], reverse=True) if n_mj > 0 else []
            # enforce DC > lowest MJ if both exist
            if n_dc > 0 and n_mj > 0 and len(mj_Eg) > 0:
                if any(dc_i <= min(mj_Eg) for dc_i in dc_Eg):
                    return 1000.0
            try:
                return -model_eff(dc_Eg, uc_Eg, mj_Eg, irradiance_np, wavelength_np)
            except Exception:
                return 1000.0
        best_eff = [0.0]
        best_x = [None]
        def callback(xk, convergence):
            eff = -objective(xk)
            if eff > best_eff[0]:
                best_eff[0] = eff
                best_x[0] = xk.copy()
        try:
            result = differential_evolution(
                objective,
                bounds,
                strategy='best1bin',
                maxiter=100,
                popsize=40,
                tol=0.001,
                seed=123981294,
                callback=callback,
                polish=True,
            )
            final_eff = -result.fun
            final_dc = sorted(result.x[:n_dc], reverse=True) if n_dc > 0 else []
            final_uc = list(result.x[n_dc:n_dc+n_uc]) if n_uc > 0 else []
            final_mj = sorted(result.x[n_dc+n_uc:], reverse=True) if n_mj > 0 else []
        except Exception:
            traceback.print_exc()
            final_eff = 0.0
            final_dc = []
            final_uc = []
            final_mj = []

        results.append({
            'n_dc': n_dc,
            'n_uc': n_uc,
            'n_mj': n_mj,
            'efficiency': final_eff,
            'dc_Eg': final_dc,
            'uc_Eg': final_uc,
            'mj_Eg': final_mj
        })

        # brief progress update
        print(f"[{idx}/{len(configs)}] {n_dc} DC, {n_uc} UC, {n_mj} MJ -> {final_eff:.4f}%")
    return results

dc_config = [1,2] #Number of down-converting layers to optimize                          
uc_config = [1,2] #Number of up-converting layers to optimize     
mj_config = [1,2] #Numebr of cells to optimize

all_results = find_optimal_configuration(dc_config, uc_config, mj_config)

rows = []
for res in all_results:
    dc_count = res['n_dc']
    uc_count = res['n_uc']
    mj_count = res['n_mj']
    eff = res['efficiency']
    dc_bg = "None"
    if dc_count > 0 and len(res['dc_Eg']) > 0:
        dc_bg = "\n".join(f"{eg:.4f}" for eg in res['dc_Eg'])
    uc_bg = "None"
    if uc_count > 0 and len(res['uc_Eg']) > 0:
        uc_bg = "\n".join(f"{eg:.4f}" for eg in res['uc_Eg'])
    mj_bg = "None"
    if mj_count > 0 and len(res['mj_Eg']) > 0:
        mj_bg = "\n".join(f"{eg:.4f}" for eg in res['mj_Eg'])
    rows.append([dc_count, uc_count, mj_count, eff, dc_bg, uc_bg, mj_bg])

df = pd.DataFrame(rows, columns=[
    "DC Count", "UC Count", "MJ Count", "Efficiency (%)",
    "DC Bandgaps (eV)", "UC Bandgaps (eV)", "MJ Bandgaps (eV)"
])

with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="results")

wb = load_workbook(OUTPUT_FILE)
ws = wb["results"]
header_map = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
cols_to_wrap = []
for col_name in ["DC Bandgaps (eV)", "UC Bandgaps (eV)", "MJ Bandgaps (eV)"]:
    if col_name in header_map:
        cols_to_wrap.append(header_map[col_name])
for col_idx in cols_to_wrap:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.alignment = Alignment(wrapText=True)
wb.save(OUTPUT_FILE)
print(f"✅ Results saved to: {OUTPUT_FILE}")
