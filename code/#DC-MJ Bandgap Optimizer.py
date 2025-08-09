#GitHub DC-MJ Bandgap Optimizer
import pandas as pd
import numpy as np
from scipy.constants import h, c, e, k
from scipy import integrate
from scipy.optimize import differential_evolution
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Alignment
file_path = r"D:\dowloads\AM1.5Global.xlsx" #Import AM1.5G Spectrum
OUTPUT_FILE = r"D:\dowloads\dcmj_results.xlsx" #Enter desired data download location
data = pd.read_excel(file_path)
wavelength = data.iloc[:, 0].values    # nm
irradiance = data.iloc[:, 1].values    # W/m^2/nm
q = e
T = 300
pi = np.pi
n = 1
CM = 1
def compute_j0(Eg):
    # Eg in eV; integrand uses same functional form you had
    prefactor = 2 * pi * q**4 / (h**3 * c**2)
    # integrand in terms of energy E (eV) as in your original code
    def integrand(E):
        # clip argument to avoid overflow
        x = np.clip(q * E / (k * T), -700, 700)
        return E**2 / (np.expm1(x))
    try:
        integral, _ = integrate.quad(integrand, Eg, np.inf, limit=200)
    except Exception:
        # If integration fails, return a large J0 to kill this candidate
        return np.inf
    return prefactor * integral
def model_eff(dc_Eg, mj_Eg, irradiance_np, wavelength_np):
    E_spectrum = h * c / (q * wavelength_np * 1e-9)  # Energy in eV
    I_dc = irradiance_np.copy()
    P_in = np.trapezoid(I_dc, wavelength_np)  # total incident power
    # Down-conversion handling (kept same)
    if len(dc_Eg) > 0:
        dλ = np.gradient(wavelength_np)
        for Eg_target in dc_Eg:
            indices_src = np.where(np.abs(E_spectrum - Eg_target) <= 0.05)[0]
            for i in indices_src:
                if I_dc[i] <= 0:
                    continue
                E_tgt = E_spectrum[i] / 2.0
                indices_dst = np.where(np.abs(E_spectrum - E_tgt) <= 0.05)[0]
                if len(indices_dst) == 0:
                    continue
                # distribute source bin power into destination bins (as you did)
                denom = np.sum(dλ[indices_dst])
                if denom <= 0:
                    continue
                I_dc[indices_dst] += I_dc[i] * dλ[i] / denom
                I_dc[i] = 0.0

    # Multijunction
    P_max_list = []
    for i, Eg_i in enumerate(mj_Eg):
        if i == 0:
            λ_max = h * c / (q * 1e-9 * Eg_i)
            mask = wavelength_np <= λ_max
        else:
            λ_min = h * c / (q * 1e-9 * mj_Eg[i-1])
            λ_max = h * c / (q * 1e-9 * Eg_i)
            mask = (wavelength_np >= λ_min) & (wavelength_np <= λ_max)

        if not np.any(mask):
            P_max_list.append(0.0)
            continue

        E_cell = h * c / (q * wavelength_np[mask] * 1e-9)
        # photon flux: integrate spectral irradiance / (q * photon energy)
        photon_flux = np.trapezoid(I_dc[mask] / (q * E_cell), wavelength_np[mask])
        J_ph = CM * q * photon_flux
        J0 = compute_j0(Eg_i)

        if J_ph <= 0 or J0 <= 0 or not np.isfinite(J0):
            P_max_list.append(0.0)
            continue

        Voc = n * k * T * np.log1p(J_ph / J0) / q
        V = np.linspace(0, Voc, 500)
        # note: using expm1 consistent with original code
        J_total = J_ph - J0 * np.expm1(V * q / (n * k * T))
        power = J_total * V
        max_power = np.max(power) if len(power) > 0 else 0.0
        P_max_list.append(max_power)

    P_collected = sum(P_max_list)
    if P_in <= 0:
        return 0.0
    return (P_collected / P_in) * 100.0
def find_optimal_configuration(x, y):
    configs = [(n_dc, n_mj) for n_dc in x for n_mj in y]
    results = []
    for n_dc, n_mj in configs:
        n_params = n_dc + n_mj
        bounds = [(0.1, 5.0)] * n_params

        def objective(xvec):
            dc_Eg = sorted(xvec[:n_dc], reverse=True) if n_dc > 0 else []
            mj_Eg = sorted(xvec[n_dc:], reverse=True) if n_mj > 0 else []
            # Ban solutions where any DC <= lowest MJ when both exist
            if n_dc > 0 and n_mj > 0:
                lowest_mj = min(mj_Eg)
                if any(dc_i <= lowest_mj for dc_i in dc_Eg):
                    return 1000.0
            try:
                return -model_eff(dc_Eg, mj_Eg, irradiance, wavelength)
            except Exception:
                return 1000.0

        best_eff = [0.0]
        best_x = [None]

        def callback(xk, convergence):
            eff = -objective(xk)
            if eff > best_eff[0]:
                best_eff[0] = eff
                best_x[0] = xk.copy()

        try:
            result = differential_evolution(
                objective,
                bounds,
                strategy='best1bin',
                maxiter=100,
                popsize=40,
                tol=0.001,
                seed=123981294,
                callback=callback,
                polish=True,
            )
            final_eff = -result.fun
            final_dc = sorted(result.x[:n_dc], reverse=True) if n_dc > 0 else []
            final_mj = sorted(result.x[n_dc:], reverse=True) if n_mj > 0 else []
        except Exception:
            # If the optimizer crashes for some config, capture the error and continue
            traceback.print_exc()
            final_eff = 0.0
            final_dc = []
            final_mj = []

        results.append({
            'n_dc': n_dc,
            'n_mj': n_mj,
            'efficiency': final_eff,
            'dc_Eg': final_dc,
            'mj_Eg': final_mj
        })

        # print a short status line (won't break output formatting)
        print(f"Finished: {n_dc} DC + {n_mj} MJ => {final_eff:.4f}%")

    return results

dc_config = [0,1,2] #Number of down-converting layers to optimize 
mj_config = [1,2,3] #Numebr of cells to optimize

all_results = find_optimal_configuration(dc_config, mj_config)
rows = []
for res in all_results:
    dc_count = res['n_dc']
    mj_count = res['n_mj']
    eff = res['efficiency']
    dc_bg = "None"
    if dc_count > 0 and len(res['dc_Eg']) > 0:
        dc_bg = "\n".join([f"{eg:.4f}" for eg in res['dc_Eg']])
    mj_bg = "None"
    if mj_count > 0 and len(res['mj_Eg']) > 0:
        mj_bg = "\n".join([f"{eg:.4f}" for eg in res['mj_Eg']])
    rows.append([dc_count, mj_count, eff, dc_bg, mj_bg])
df = pd.DataFrame(rows, columns=["DC Count", "MJ Count", "Efficiency (%)", "DC Bandgaps (eV)", "MJ Bandgaps (eV)"])
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="results")
wb = load_workbook(OUTPUT_FILE)
ws = wb["results"]
header_map = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
cols_to_wrap = []
for col_name in ["DC Bandgaps (eV)", "MJ Bandgaps (eV)"]:
    if col_name in header_map:
        cols_to_wrap.append(header_map[col_name])
for col_idx in cols_to_wrap:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.alignment = Alignment(wrapText=True)
wb.save(OUTPUT_FILE)
print(f"✅ Results saved to: {OUTPUT_FILE}")
